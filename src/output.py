import os.path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import *

import config

"""
{
'customer_data': 
    {
    'first_name': 'Sam', 
    'last_name': 'Smith', 
    'email': 'test@gmail.com', 
    'phone_number': '1231231234', 
    'address': '100 Test Rd', 
    'apartment': '', 
    'city': 'RealCity', 
    'state': 'Alabama', 
    'country': 'United States', 
    'zip': '12345'
    },
'order_data': [
        {
        'item': 'Dress', 
        'item_details': 'A small orange dress', 
        'size': 'Small', 
        'notes': 'Split sides'
        }, 
        {'item': 'Dress', 
        'item_details': '', 
        'size': 'XXS', 
        'notes': ''
        }
    ]
}
"""


def add_header(workbook):
    sheet = workbook.active
    header = ["First Name", "Last Name", "Email", "Phone Number", "Address", "Apartment", "City", "State",
              "Country", "Zip", "Item", "Color", "Item Details", "Size", "Notes", "Multiple Orders"]
    my_color = colors.Color(rgb='00C0C0C0')
    my_fill = fills.PatternFill(patternType='solid', fgColor=my_color)

    index = -1
    for col in sheet.iter_cols(min_row=1, max_col=len(header), max_row=1):
        for cell in col:
            index += 1
            cell.value = header[index]
            cell.fill = my_fill


class Output:
    def __init__(self, file_path, file_name, data):
        # file = file_path + "/" + file_name + ".xlsx"
        file = config.values["output_directory"] + "/" + config.values["output_file_name"] + ".xlsx"

        if not os.path.exists(file):
            workbook = Workbook()

        else:
            workbook = load_workbook(file)

        add_header(workbook)

        worksheet = workbook.active

        customer_details = []
        orders = []

        for key in data.keys():
            if key != "order_data" and key != "multiple_orders":
                for customer_detail in data[key].values():
                    print("\tCUSTOMER DETAIL: " + str(customer_detail))
                    customer_details.append(customer_detail)
            elif key == "order_data":
                for order in data[key]:
                    o = []
                    for order_details in order.values():
                        o.append(order_details)

                    orders.append(o)

        for order in orders:
            line = customer_details + order

            if len(orders) > 1:
                line += ["Yes"]
            else:
                line += ["No"]

            worksheet.append(line)



        workbook.save(file)
